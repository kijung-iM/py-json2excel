import json
from openpyxl import load_workbook

wb = load_workbook("agent-options.xlsx")
wsht = wb["agentsetting"]

with open("agentsettings.json", "r") as json_file:
    agentsettings = json.load(json_file)

row = 2

for i in agentsettings:
    options = agentsettings[i]
    for j in options:
        if j == "key":
            print(j, " - ", options[j])
            wsht.cell(row, 5).value = options[j]
        elif j == "group":
            # loop -> 차례로 입력
            print("group:")
            groups = options[j]
            gcol = 3
            for k in groups:
                print("\t", k)
                wsht.cell(row, gcol).value = k
                gcol += 1
        elif j == "description":
            # loop -> 차례로 입력
            languages = options[j]
            print("languages:")
            lcol = 12
            for lang in languages:
                print(lang, " - ", languages[lang])
                wsht.cell(row, lcol).value = languages[lang]
                lcol += 1
        elif j == "badge":
            # loop -> 차례로 입력
            print("badge:")
            badges = options[j]
            for badge in badges:
                print("\t", badge, " : ", badges[badge])
                wsht.cell(row, 15).value = badge
                wsht.cell(row, 16).value = badges[badge]
        elif j == "defaultValue":
            print(j, " - ", options[j])
            wsht.cell(row, 17).value = options[j]
        elif j == "type":
            print(j, " - ", options[j])
            wsht.cell(row, 18).value = options[j]
        elif j == "unit":
            print(j, " - ", options[j])
            wsht.cell(row, 19).value = options[j]
    row += 1
    print("==================================")

wb.save('test.xlsx')